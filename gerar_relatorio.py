"""
Script para gerar relatório Excel completo com dados reais.
"""
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from src.config.settings import MESES_PT, LISTA_PRODUTOS
from src.data_processing.loader import (
    carregar_e_processar_dados,
)
from src.reports.tabela_produtos_horizontal import (
    criar_tabela_produtos_horizontal,
)
from src.reports.tabela_produto_individual import (
    criar_tabela_produto_individual,
)
from src.reports.relatorio_mix import criar_relatorio_mix
from src.reports.resumo_geral import criar_resumo_geral


def ajustar_largura_colunas(ws):
    """Ajusta largura das colunas automaticamente."""
    from openpyxl.cell.cell import MergedCell
    for column in ws.columns:
        max_length = 0
        col_letter = None
        for cell in column:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                if (
                    cell.value
                    and len(str(cell.value)) > max_length
                ):
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                continue
        if col_letter is None:
            continue
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width


def criar_aba_resumo(wb, df, mes, ano):
    """Cria aba de resumo executivo."""
    print(f"\nCriando aba: Resumo Executivo...")
    
    ws = wb.create_sheet("Resumo Executivo", 0)
    
    mes_nome_pt = MESES_PT[mes]
    
    ws['A1'] = f"RELATÓRIO DE VENDAS - {mes_nome_pt.upper()}/{ano}"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('A1:D1')
    
    row = 3
    ws[f'A{row}'] = "RESUMO GERAL"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    
    row += 1
    ws[f'A{row}'] = "Total de Vendas (VLR BASE):"
    ws[f'B{row}'] = df['VALOR'].sum()
    ws[f'B{row}'].number_format = 'R$ #,##0.00'
    
    row += 1
    ws[f'A{row}'] = "Total de Pontos:"
    ws[f'B{row}'] = df['pontos'].sum()
    ws[f'B{row}'].number_format = '#,##0'
    
    row += 1
    ws[f'A{row}'] = "Total de Transações:"
    ws[f'B{row}'] = len(df)
    
    row += 1
    ws[f'A{row}'] = "Ticket Médio:"
    ws[f'B{row}'] = df['VALOR'].mean()
    ws[f'B{row}'].number_format = 'R$ #,##0.00'
    
    row += 1
    ws[f'A{row}'] = "Número de Consultores:"
    ws[f'B{row}'] = df['CONSULTOR'].nunique()
    
    row += 1
    ws[f'A{row}'] = "Número de Lojas:"
    ws[f'B{row}'] = df['LOJA'].nunique()
    
    row += 1
    ws[f'A{row}'] = "Número de Regiões:"
    ws[f'B{row}'] = df['REGIAO'].nunique()
    
    if 'SUBTIPO' in df.columns:
        row += 2
        ws[f'A{row}'] = "PRODUTOS ESPECIAIS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        ws[f'A{row}'] = "Super Conta:"
        ws[f'B{row}'] = len(df[df['SUBTIPO'] == 'SUPER CONTA'])
        
        row += 1
        ws[f'A{row}'] = "Emissões:"
        ws[f'B{row}'] = len(df[df['SUBTIPO'] == 'EMISSAO'])
        
        row += 1
        ws[f'A{row}'] = "FGTS:"
        ws[f'B{row}'] = len(df[df['SUBTIPO'] == 'FGTS'])
    
    ajustar_largura_colunas(ws)
    
    print(f"   ✓ Resumo criado")


def criar_aba_ranking_consultores(wb, df):
    """Cria aba com ranking de consultores."""
    print(f"\nCriando aba: Ranking Consultores...")
    
    ws = wb.create_sheet("Ranking Consultores")
    
    ranking = df.groupby('CONSULTOR').agg({
        'pontos': 'sum',
        'VALOR': 'sum',
        'LOJA': 'first',
        'REGIAO': 'first'
    }).reset_index()
    
    ranking = ranking.sort_values('pontos', ascending=False)
    ranking['posicao'] = range(1, len(ranking) + 1)
    ranking = ranking[['posicao', 'CONSULTOR', 'LOJA', 'REGIAO', 'pontos', 'VALOR']]
    
    headers = ['Posição', 'Consultor', 'Loja', 'Região', 'Pontos', 'Valor']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, row_data in enumerate(ranking.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    for row in range(2, len(ranking) + 2):
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=6).number_format = 'R$ #,##0.00'
    
    ajustar_largura_colunas(ws)
    
    print(f"   ✓ Ranking de {len(ranking)} consultores criado")


def criar_aba_ranking_lojas(wb, df):
    """Cria aba com ranking de lojas."""
    print(f"\nCriando aba: Ranking Lojas...")
    
    ws = wb.create_sheet("Ranking Lojas")
    
    ranking = df.groupby('LOJA').agg({
        'pontos': 'sum',
        'VALOR': 'sum',
        'CONSULTOR': 'nunique',
        'REGIAO': 'first'
    }).reset_index()
    
    ranking.columns = ['LOJA', 'pontos', 'VALOR', 'num_consultores', 'REGIAO']
    ranking = ranking.sort_values('pontos', ascending=False)
    ranking['posicao'] = range(1, len(ranking) + 1)
    ranking = ranking[['posicao', 'LOJA', 'REGIAO', 'num_consultores', 'pontos', 'VALOR']]
    
    headers = ['Posição', 'Loja', 'Região', 'Consultores', 'Pontos', 'Valor']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, row_data in enumerate(ranking.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    for row in range(2, len(ranking) + 2):
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=6).number_format = 'R$ #,##0.00'
    
    ajustar_largura_colunas(ws)
    
    print(f"   ✓ Ranking de {len(ranking)} lojas criado")


def criar_aba_por_regiao(wb, df):
    """Cria aba com análise por região."""
    print(f"\nCriando aba: Por Região...")
    
    ws = wb.create_sheet("Por Região")
    
    analise = df.groupby('REGIAO').agg({
        'pontos': 'sum',
        'VALOR': 'sum',
        'LOJA': 'nunique',
        'CONSULTOR': 'nunique'
    }).reset_index()
    
    analise.columns = ['REGIAO', 'pontos', 'VALOR', 'num_lojas', 'num_consultores']
    analise = analise.sort_values('pontos', ascending=False)
    
    headers = ['Região', 'Pontos', 'Valor', 'Lojas', 'Consultores']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, row_data in enumerate(analise.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    for row in range(2, len(analise) + 2):
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).number_format = 'R$ #,##0.00'
    
    ajustar_largura_colunas(ws)
    
    print(f"   ✓ Análise de {len(analise)} regiões criada")


def criar_aba_tabela_produtos(wb, df, df_metas, mes, ano, dia_atual=None):
    """Cria aba com produtos por loja em formato horizontal, agrupados por região."""
    print(f"\nCriando aba: Produtos por Loja...")
    
    ws = wb.create_sheet("Produtos por Loja")
    
    tabelas_por_regiao = criar_tabela_produtos_horizontal(df, df_metas, ano, mes, dia_atual)
    
    current_row = 1
    total_linhas = 0
    
    for regiao, tabela in sorted(tabelas_por_regiao.items()):
        region_cell = ws.cell(row=current_row, column=1, value=f"REGIÃO: {regiao}")
        region_cell.font = Font(bold=True, size=12, color="FFFFFF")
        region_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        region_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        num_colunas = len(tabela.columns)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_colunas)
        current_row += 1
        
        headers = list(tabela.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        current_row += 1
        
        for _, row_data in tabela.iterrows():
            is_total_row = str(row_data['Loja']).startswith('TOTAL ')
            
            for col_idx, (col_name, value) in enumerate(zip(headers, row_data), 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                
                if is_total_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                if 'Meta' in col_name and col_name != 'Loja':
                    cell.number_format = '#,##0'
                elif any(x in col_name for x in ['CNC', 'SAQUE', 'CLT', 'CONSIGNADO', 'PACK', 'Projeção', 'Média DU', 'Ticket Médio']) and 'Loja' not in col_name and '%' not in col_name:
                    cell.number_format = 'R$ #,##0.00'
                elif '%' in col_name:
                    cell.number_format = '0.00"%"'
            
            current_row += 1
            total_linhas += 1
        
        current_row += 1
    
    current_row += 1
    nota_cell = ws.cell(row=current_row, column=1, value="* PACK = FGTS, ANT. BEN. e CNC 13º")
    nota_cell.font = Font(italic=True, size=9, color="666666")
    nota_cell.alignment = Alignment(horizontal="left")
    
    ajustar_largura_colunas(ws)
    
    print(f"   ✓ Tabela de produtos por loja criada ({total_linhas} lojas)")


def criar_aba_produto_individual(wb, df, df_metas, produto_nome, ano, mes, dia_atual=None):
    """Cria aba individual para um produto específico."""
    print(f"\nCriando aba: {produto_nome}...")
    
    ws = wb.create_sheet(produto_nome)
    
    tabelas_por_regiao = criar_tabela_produto_individual(df, df_metas, produto_nome, ano, mes, dia_atual)
    
    current_row = 1
    total_linhas = 0
    
    for regiao, tabela in sorted(tabelas_por_regiao.items()):
        region_cell = ws.cell(row=current_row, column=1, value=f"REGIÃO: {regiao}")
        region_cell.font = Font(bold=True, size=12, color="FFFFFF")
        region_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        region_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        num_colunas = len(tabela.columns)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_colunas)
        current_row += 1
        
        headers = list(tabela.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        current_row += 1
        
        for _, row_data in tabela.iterrows():
            is_total_row = str(row_data['Loja']).startswith('TOTAL ')
            
            for col_idx, (col_name, value) in enumerate(zip(headers, row_data), 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                
                if is_total_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                if 'Meta' in col_name and col_name != 'Loja':
                    cell.number_format = 'R$ #,##0.00'
                elif any(x in col_name for x in ['Valor', 'Média DU', 'Ticket Médio', 'Projeção']) and 'Loja' not in col_name:
                    cell.number_format = 'R$ #,##0.00'
                elif '%' in col_name:
                    cell.number_format = '0.00"%"'
            
            current_row += 1
            total_linhas += 1
        
        current_row += 1
    
    if produto_nome == 'PACK':
        current_row += 1
        nota_cell = ws.cell(row=current_row, column=1, value="* PACK = FGTS, ANT. BEN. e CNC 13º")
        nota_cell.font = Font(italic=True, size=9, color="666666")
        nota_cell.alignment = Alignment(horizontal="left")
    
    ajustar_largura_colunas(ws)
    
    print(f"   ✓ Relatório {produto_nome} criado ({total_linhas} lojas)")


def criar_aba_tabela_consultores(wb, df, df_metas, mes, ano, df_supervisores=None):
    """Cria aba com tabela detalhada de consultores."""
    print(f"\nCriando aba: Tabela Consultores...")
    
    ws = wb.create_sheet("Tabela Consultores")
    
    tabela = criar_tabela_por_consultor(df, df_metas, ano, mes, df_supervisores=df_supervisores)
    
    headers = list(tabela.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, row_data in enumerate(tabela.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    for row in range(2, len(tabela) + 2):
        ws.cell(row=row, column=3).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=6).number_format = '0.00"%"'
        ws.cell(row=row, column=7).number_format = '#,##0'
        ws.cell(row=row, column=8).number_format = '0.00"%"'
        ws.cell(row=row, column=9).number_format = '#,##0.00'
        ws.cell(row=row, column=10).number_format = '#,##0.00'
    
    ajustar_largura_colunas(ws)
    
    print(f"   ✓ Tabela de {len(tabela)} consultores criada")


def criar_aba_tabela_lojas(wb, df, df_metas, mes, ano):
    """Cria aba com tabela detalhada de lojas."""
    print(f"\nCriando aba: Tabela Lojas...")
    
    ws = wb.create_sheet("Tabela Lojas")
    
    tabela = criar_tabela_por_loja(df, df_metas, ano, mes)
    
    headers = list(tabela.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, row_data in enumerate(tabela.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    for row in range(2, len(tabela) + 2):
        ws.cell(row=row, column=2).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=5).number_format = '0.00"%"'
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=7).number_format = '0.00"%"'
        ws.cell(row=row, column=8).number_format = '#,##0.00'
        ws.cell(row=row, column=9).number_format = '#,##0.00'
    
    ajustar_largura_colunas(ws)
    
    print(f"   ✓ Tabela de {len(tabela)} lojas criada")


def criar_aba_relatorio_mix(wb, df, df_metas, mes, ano, dia_atual=None):
    """Cria aba com relatório MIX de produtos."""
    print(f"\nCriando aba: Relatório MIX...")
    
    ws = wb.create_sheet("Relatório MIX")
    
    tabela = criar_relatorio_mix(df, df_metas, ano, mes, dia_atual)
    
    headers = list(tabela.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, row_data in enumerate(tabela.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            if row_idx == len(tabela) + 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    for row in range(2, len(tabela) + 2):
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=6).number_format = '0.00"%"'
        ws.cell(row=row, column=7).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=8).number_format = '#,##0.00'
        ws.cell(row=row, column=9).number_format = '#,##0.00'
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    print(f"   ✓ Relatório MIX criado ({len(tabela)-1} produtos)")


def criar_aba_resumo_geral(wb, df, df_metas, mes, ano, dia_atual=None, df_supervisores=None):
    """Cria aba com resumo geral consolidado."""
    print(f"\nCriando aba: Resumo Geral...")
    
    ws = wb.create_sheet("Resumo Geral", 0)
    
    resumo = criar_resumo_geral(df, df_metas, ano, mes, dia_atual, df_supervisores)
    
    row_atual = 1
    
    ws.merge_cells(f'A{row_atual}:B{row_atual}')
    titulo = ws.cell(row=row_atual, column=1, value="RESUMO GERAL - CONSOLIDADO")
    titulo.font = Font(bold=True, size=14, color="FFFFFF")
    titulo.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="TOTAIS GERAIS").font = Font(bold=True, size=12)
    row_atual += 1
    
    for _, row_data in resumo['totais_gerais'].iterrows():
        ws.cell(row=row_atual, column=1, value=row_data['Métrica']).font = Font(bold=True)
        cell_valor = ws.cell(row=row_atual, column=2, value=row_data['Valor'])
        
        if 'R$' in row_data['Métrica'] or 'Valor' in row_data['Métrica']:
            cell_valor.number_format = 'R$ #,##0.00'
        elif '%' in row_data['Métrica']:
            cell_valor.number_format = '0.00"%"'
        elif 'Quantidade' in row_data['Métrica'] or 'Pontos' in row_data['Métrica'] or 'Meta' in row_data['Métrica']:
            cell_valor.number_format = '#,##0'
        else:
            cell_valor.number_format = '#,##0.00'
        
        row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="RESUMO POR REGIÃO").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['por_regiao'].empty:
        headers_regiao = list(resumo['por_regiao'].columns)
        for col_idx, header in enumerate(headers_regiao, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['por_regiao'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                
                if col_idx == 3:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx in [2, 4, 5]:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = '0.00"%"'
                elif col_idx == 7:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 8:
                    cell.number_format = '#,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    # ========== TOP 10 LOJAS ==========
    
    ws.cell(row=row_atual, column=1, value="TOP 10 LOJAS - POR PONTOS").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['top_lojas'].empty:
        headers = list(resumo['top_lojas'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['top_lojas'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="TOP 10 LOJAS - POR TICKET MÉDIO").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['top_lojas_ticket_medio'].empty:
        headers = list(resumo['top_lojas_ticket_medio'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['top_lojas_ticket_medio'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="TOP 10 LOJAS - POR MÉDIA DU").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['top_lojas_media_du'].empty:
        headers = list(resumo['top_lojas_media_du'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['top_lojas_media_du'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = '#,##0.00'
                elif col_idx == 7:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="TOP 10 LOJAS - POR PRODUTO").font = Font(bold=True, size=12)
    row_atual += 1
    
    for produto, ranking in resumo['top_lojas_por_produto'].items():
        if ranking.empty:
            continue
        
        ws.cell(row=row_atual, column=1, value=f"{produto}").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row_atual, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        row_atual += 1
        
        headers = list(ranking.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in ranking.iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 7:
                    cell.number_format = '#,##0.00'
            row_atual += 1
        
        row_atual += 1
    
    row_atual += 1
    
    ws.cell(row=row_atual, column=1, value="TOP 10 LOJAS - POR ATINGIMENTO META PRATA").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['top_lojas_atingimento'].empty:
        headers = list(resumo['top_lojas_atingimento'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['top_lojas_atingimento'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = '#,##0'
                elif col_idx == 7:
                    cell.number_format = '0.00"%"'
                elif col_idx == 8:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    # ========== TOP 10 CONSULTORES ==========
    
    ws.cell(row=row_atual, column=1, value="TOP 10 CONSULTORES - POR PONTOS").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['top_consultores'].empty:
        headers = list(resumo['top_consultores'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['top_consultores'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 4:
                    cell.number_format = '#,##0'
                elif col_idx == 5:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 6:
                    cell.number_format = '#,##0'
                elif col_idx == 7:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="TOP 10 CONSULTORES - POR TICKET MÉDIO").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['top_consultores_ticket_medio'].empty:
        headers = list(resumo['top_consultores_ticket_medio'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['top_consultores_ticket_medio'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 4:
                    cell.number_format = '#,##0'
                elif col_idx == 5:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 6:
                    cell.number_format = '#,##0'
                elif col_idx == 7:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="TOP 10 CONSULTORES - POR MÉDIA DU").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['top_consultores_media_du'].empty:
        headers = list(resumo['top_consultores_media_du'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['top_consultores_media_du'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 4:
                    cell.number_format = '#,##0'
                elif col_idx == 5:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 6:
                    cell.number_format = '#,##0'
                elif col_idx == 7:
                    cell.number_format = '#,##0.00'
                elif col_idx == 8:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="TOP 10 CONSULTORES - POR PRODUTO").font = Font(bold=True, size=12)
    row_atual += 1
    
    for produto, ranking in resumo['top_consultores_por_produto'].items():
        if ranking.empty:
            continue
        
        ws.cell(row=row_atual, column=1, value=f"{produto}").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row_atual, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        row_atual += 1
        
        headers = list(ranking.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in ranking.iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 4:
                    cell.number_format = '#,##0'
                elif col_idx == 5:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 6:
                    cell.number_format = '#,##0'
                elif col_idx == 7:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 8:
                    cell.number_format = '#,##0.00'
            row_atual += 1
        
        row_atual += 1
    
    row_atual += 1
    
    ws.cell(row=row_atual, column=1, value="TOP 10 CONSULTORES - POR ATINGIMENTO META PRATA").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['top_consultores_atingimento'].empty:
        headers = list(resumo['top_consultores_atingimento'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['top_consultores_atingimento'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 4:
                    cell.number_format = '#,##0'
                elif col_idx == 5:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 6:
                    cell.number_format = '#,##0'
                elif col_idx == 7:
                    cell.number_format = '#,##0'
                elif col_idx == 8:
                    cell.number_format = '0.00"%"'
                elif col_idx == 9:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    # ========== RANKING DE REGIÃO ==========
    
    ws.cell(row=row_atual, column=1, value="RANKING DE REGIÃO - POR PONTOS").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['ranking_regiao_pontos'].empty:
        headers = list(resumo['ranking_regiao_pontos'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['ranking_regiao_pontos'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="RANKING DE REGIÃO - POR TICKET MÉDIO").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['ranking_regiao_ticket_medio'].empty:
        headers = list(resumo['ranking_regiao_ticket_medio'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['ranking_regiao_ticket_medio'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="RANKING DE REGIÃO - POR MÉDIA DU").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['ranking_regiao_media_du'].empty:
        headers = list(resumo['ranking_regiao_media_du'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['ranking_regiao_media_du'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx in [6, 7]:
                    cell.number_format = '#,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="RANKING DE REGIÃO - POR PRODUTO").font = Font(bold=True, size=12)
    row_atual += 1
    
    for produto, ranking in resumo['ranking_regiao_por_produto'].items():
        if ranking.empty:
            continue
        
        ws.cell(row=row_atual, column=1, value=f"{produto}").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row_atual, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        row_atual += 1
        
        headers = list(ranking.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in ranking.iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx in [6, 7]:
                    cell.number_format = '#,##0.00'
            row_atual += 1
        
        row_atual += 1
    
    row_atual += 1
    
    ws.cell(row=row_atual, column=1, value="RANKING DE REGIÃO - POR ATINGIMENTO META PRATA").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['ranking_regiao_atingimento'].empty:
        headers = list(resumo['ranking_regiao_atingimento'].columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['ranking_regiao_atingimento'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = '#,##0'
                elif col_idx == 7:
                    cell.number_format = '0.00"%"'
                elif col_idx == 8:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ajustar_largura_colunas(ws)
    
    print(f"   ✓ Resumo Geral criado")


def gerar_relatorio(mes=3, ano=2026):
    """Gera relatório Excel completo."""
    print(f"\n{'#'*80}")
    print(f"# GERAÇÃO DE RELATÓRIO EXCEL")
    print(f"# Período: {mes:02d}/{ano}")
    print(f"{'#'*80}")
    
    df, df_metas, df_supervisores, dia_atual_para_du = carregar_e_processar_dados(mes, ano)
    
    Path('outputs/relatorios_excel').mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo = f'outputs/relatorios_excel/relatorio_vendas_{mes:02d}_{ano}_{timestamp}.xlsx'
    
    print(f"\n{'='*80}")
    print(f"CRIANDO ARQUIVO EXCEL")
    print(f"{'='*80}")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    criar_aba_resumo_geral(wb, df, df_metas, mes, ano, dia_atual_para_du, df_supervisores)
    criar_aba_resumo(wb, df, mes, ano)
    criar_aba_tabela_produtos(wb, df, df_metas, mes, ano, dia_atual_para_du)
    
    for produto in LISTA_PRODUTOS:
        criar_aba_produto_individual(
            wb, df, df_metas, produto,
            ano, mes, dia_atual_para_du
        )
    
    criar_aba_relatorio_mix(wb, df, df_metas, mes, ano, dia_atual_para_du)
    
    print(f"\n{'='*80}")
    print(f"SALVANDO ARQUIVO")
    print(f"{'='*80}")
    print(f"\nArquivo: {arquivo}")
    
    wb.save(arquivo)
    
    print(f"\n{'#'*80}")
    print(f"# ✓ RELATÓRIO GERADO COM SUCESSO!")
    print(f"# Arquivo: {arquivo}")
    print(f"# Total de pontos: {df['pontos'].sum():,.0f}")
    print(f"# Total de vendas: R$ {df['VALOR'].sum():,.2f}")
    print(f"{'#'*80}\n")
    
    return arquivo


if __name__ == "__main__":
    arquivo = gerar_relatorio(mes=3, ano=2026)
