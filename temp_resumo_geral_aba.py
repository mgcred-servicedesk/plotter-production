from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def criar_aba_resumo_geral(wb, df, df_metas, mes, ano, dia_atual=None):
    """Cria aba com resumo geral consolidado."""
    from src.reports.resumo_geral import criar_resumo_geral
    
    print(f"\nCriando aba: Resumo Geral...")
    
    ws = wb.create_sheet("Resumo Geral", 0)
    
    resumo = criar_resumo_geral(df, df_metas, ano, mes, dia_atual)
    
    row_atual = 1
    
    ws.merge_cells(f'A{row_atual}:B{row_atual}')
    titulo = ws.cell(row=row_atual, column=1, value="RESUMO GERAL - CONSOLIDADO")
    titulo.font = Font(bold=True, size=14, color="FFFFFF")
    titulo.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="TOTAIS GERAIS").font = Font(bold=True, size=12)
    row_atual += 1
    
    for _, row_data in resumo['totais_gerais'].iterrows():
        ws.cell(row=row_atual, column=1, value=row_data['Métrica']).font = Font(bold=True)
        cell_valor = ws.cell(row=row_atual, column=2, value=row_data['Valor'])
        
        if 'R$' in row_data['Métrica'] or 'Valor' in row_data['Métrica']:
            cell_valor.number_format = 'R$ #,##0.00'
        elif '%' in row_data['Métrica']:
            cell_valor.number_format = '0.00"%"'
        elif 'Quantidade' in row_data['Métrica'] or 'Pontos' in row_data['Métrica'] or 'Meta' in row_data['Métrica']:
            cell_valor.number_format = '#,##0'
        else:
            cell_valor.number_format = '#,##0.00'
        
        row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="RESUMO POR REGIÃO").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['por_regiao'].empty:
        headers_regiao = list(resumo['por_regiao'].columns)
        for col_idx, header in enumerate(headers_regiao, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['por_regiao'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                
                if col_idx == 3:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx in [2, 4, 5]:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = '0.00"%"'
                elif col_idx == 7:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 8:
                    cell.number_format = '#,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="TOP 10 LOJAS").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['top_lojas'].empty:
        headers_lojas = list(resumo['top_lojas'].columns)
        for col_idx, header in enumerate(headers_lojas, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['top_lojas'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="TOP 10 CONSULTORES").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['top_consultores'].empty:
        headers_consultores = list(resumo['top_consultores'].columns)
        for col_idx, header in enumerate(headers_consultores, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['top_consultores'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                
                if col_idx == 1:
                    cell.number_format = '0'
                elif col_idx == 3:
                    cell.number_format = '#,##0'
                elif col_idx == 4:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx == 5:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = 'R$ #,##0.00'
            row_atual += 1
    
    row_atual += 2
    
    ws.cell(row=row_atual, column=1, value="PRODUTOS MIX").font = Font(bold=True, size=12)
    row_atual += 1
    
    if not resumo['produtos_mix'].empty:
        headers_mix = list(resumo['produtos_mix'].columns)
        for col_idx, header in enumerate(headers_mix, 1):
            cell = ws.cell(row=row_atual, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        row_atual += 1
        
        for _, row_data in resumo['produtos_mix'].iterrows():
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_atual, column=col_idx, value=value)
                
                if col_idx == 2:
                    cell.number_format = '#,##0'
                elif col_idx == 3:
                    cell.number_format = 'R$ #,##0.00'
                elif col_idx in [4, 5]:
                    cell.number_format = '#,##0'
                elif col_idx == 6:
                    cell.number_format = '0.00"%"'
                elif col_idx in [7, 8]:
                    cell.number_format = '#,##0.00'
            row_atual += 1
    
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        max_length = 0
        try:
            for cell in ws[col_letter]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        except:
            pass
    
    print(f"   ✓ Resumo Geral criado")
